from flask import Flask
from flask import request
from openpyxl import Workbook
from flask_cors import CORS
import pyqrcode
import openpyxl
wb = openpyxl.load_workbook('Invoice.xlsx')

# grab the active worksheet
ws = wb.active


app = Flask(__name__)
CORS(app)

@app.route("/")
def hello():
    return "Hello World!s"

@app.route("/GetExcelData", methods=['POST'])
def hello2():
    try:
        import json
        req_data = request.get_json()
        generate_qr()
        ws['A1'] = "Invoice No. "
        ws['B1'] = req_data['VoucherId']
        ws['D6'] = "Party: "+req_data['Account']
        ws['G38'] = "Rs. "+str(req_data['amount'])
        ws['A44'] = req_data['narration']
        ws['G1'] = req_data['mydate']
        j = 8
        for i, v in enumerate(req_data['children']):
            j = j+1
            print('B'+str(j))
            ws['A'+str(j)] = str(i+1)
            ws['B'+str(j)] = v['ledgername']
            ws['G'+str(j)] = v['Amount']
        img = openpyxl.drawing.image.Image('url.png')
        img.anchor = 'B22'
        ws.add_image(img)
        wb.save("upload/"+req_data['VoucherId']+".xlsx")
        return req_data
    except Exception as e: 
        return str(e)
        


@app.route("/Tally", methods=['POST'])
def hello3():
    data = request.stream.read()
    url = "http://localhost:9000"
    headers = {
        'Content-Type': "text/xml",
    }
    import requests
    response = requests.request("POST", url, data=data, headers=headers)
    return response.text

def generate_qr():
    link_to_post = "https://medium.com/@ngengesenior/qr-codes-generation-with-python-377735be6c5f"
    url = pyqrcode.create(link_to_post)
    url.png('url.png', scale=5)
    print("Printing QR code")


if __name__ == "__main__":
    app.run('0.0.0.0')
